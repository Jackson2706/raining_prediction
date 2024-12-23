import pandas as pd
import openpyxl
from openpyxl.styles import Font

# Read CSV file into DataFrame
df = pd.read_csv('/home/jackson-devworks/Desktop/raining_prediction/new_data_with_predictions_ThacXang.csv')
output_filename = '/home/jackson-devworks/Desktop/raining_prediction/output/model_comparison_results_ThacXang.xlsx'
# Replace these column names with the actual new names you have in your DataFrame
rainfall_col = 'Rainfall'
ensemble_col = 'Ensemble_predictions'
rf_col = 'RandomForestRegressor_predictions'
xgb_col = 'BaggingRegressor_pred'
bagging_col = 'XGBRegressor_pred'
lgbm_col = 'LGBMRegressor_pred'
histgb_col = 'HistGradientBoostingRegressor_pred'

reference_model_list = ["COMS","GFS","MITSUISHI2011_D02","LING3_D02","LINKF_D02","LINBMJ_D02","ETAKF_D02","ETAG3_D02","ETABMJ_D02"]
window_sizes = [i for i in range(1, 73) if i % 24 == 0]

# List to store results
results = []

for window_size in window_sizes:
    # Create rolling window sums for prediction columns
    df['Rainfall_sum'] = df[rainfall_col].rolling(window=window_size).sum()
    df['Ensemble_predictions_sum'] = df[ensemble_col].rolling(window=window_size).sum()
    df['RandomForestRegressor_predictions_sum'] = df[rf_col].rolling(window=window_size).sum()
    df['XGBRegressor_pred_sum'] = df[xgb_col].rolling(window=window_size).sum()
    df['BaggingRegressor_pred_sum'] = df[bagging_col].rolling(window=window_size).sum()
    df['LGBMRegressor_pred_sum'] = df[lgbm_col].rolling(window=window_size).sum()
    df['HistGradientBoostingRegressor_pred_sum'] = df[histgb_col].rolling(window=window_size).sum()
    for model_name in reference_model_list:
        df[f'{model_name}_sum'] = df[model_name].rolling(window=window_size).sum()
    # Filter rows based on Rainfall_sum condition
    df = df[df['Rainfall_sum'] >= 10]
    # Calculate the new metric
    df['Ensemble_metric'] = df.apply(
        lambda row: abs(row['Ensemble_predictions_sum'] - row['Rainfall_sum']) /
                    max(row['Ensemble_predictions_sum'], row['Rainfall_sum']) if max(row['Ensemble_predictions_sum'], row['Rainfall_sum']) > 0 else 0, axis=1)
    df['RandomForestRegressor_metric'] = df.apply(
        lambda row: abs(row['RandomForestRegressor_predictions_sum'] - row['Rainfall_sum']) /
                    max(row['RandomForestRegressor_predictions_sum'], row['Rainfall_sum']) if max(row['RandomForestRegressor_predictions_sum'], row['Rainfall_sum']) > 0 else 0, axis=1)
    df['XGBRegressor_metric'] = df.apply(
        lambda row: abs(row['XGBRegressor_pred_sum'] - row['Rainfall_sum']) /
                    max(row['XGBRegressor_pred_sum'], row['Rainfall_sum']) if max(row['XGBRegressor_pred_sum'], row['Rainfall_sum']) > 0 else 0, axis=1)
    df['BaggingRegressor_metric'] = df.apply(
        lambda row: abs(row['BaggingRegressor_pred_sum'] - row['Rainfall_sum']) /
                    max(row['BaggingRegressor_pred_sum'], row['Rainfall_sum']) if max(row['BaggingRegressor_pred_sum'], row['Rainfall_sum']) > 0 else 0, axis=1)
    df['LGBMRegressor_metric'] = df.apply(
        lambda row: abs(row['LGBMRegressor_pred_sum'] - row['Rainfall_sum']) /
                    max(row['LGBMRegressor_pred_sum'], row['Rainfall_sum']) if max(row['LGBMRegressor_pred_sum'], row['Rainfall_sum']) > 0 else 0, axis=1)
    df['HistGradientBoostingRegressor_metric'] = df.apply(
        lambda row: abs(row['HistGradientBoostingRegressor_pred_sum'] - row['Rainfall_sum']) /
                    max(row['HistGradientBoostingRegressor_pred_sum'], row['Rainfall_sum']) if max(row['HistGradientBoostingRegressor_pred_sum'], row['Rainfall_sum']) > 0 else 0, axis=1)
    for model_name in reference_model_list:
        df[f'{model_name}_metric'] = df.apply(
            lambda row: abs(row[f'{model_name}_sum'] - row['Rainfall_sum']) /
                        max(row[f'{model_name}_sum'], row['Rainfall_sum']) if max(row[f'{model_name}_sum'], row['Rainfall_sum']) > 0 else 0, axis=1)
    

    # Calculate the average metric for each model
    ours_avg_metric = df['Ensemble_metric'].mean()
    mlp_avg_metric = df['RandomForestRegressor_metric'].mean()
    omp_cv_avg_metric = df['XGBRegressor_metric'].mean()
    bayesian_ridge_avg_metric = df['BaggingRegressor_metric'].mean()
    lasso_lars_ic_avg_metric = df['LGBMRegressor_metric'].mean()
    lars_cv_avg_metric = df['HistGradientBoostingRegressor_metric'].mean()

    # Store results in the list
    result_item = {
        'window_size': window_size,
        'Ensemble_avg_metric': ours_avg_metric,
        'RandomForestRegressor_avg_metric': mlp_avg_metric,
        'XGBRegressor_avg_metric': omp_cv_avg_metric,
        'BaggingRegressor_avg_metric': bayesian_ridge_avg_metric,
        'LGBMRegressor_avg_metric': lasso_lars_ic_avg_metric,
        'HistGradientBoostingRegressor_avg_metric': lars_cv_avg_metric,
    }
    for model_name in reference_model_list:
        val = df[f'{model_name}_metric'].mean()
        result_item[f"{model_name}_avg_metric"] = val

    results.append(result_item)
# Convert the results list to a DataFrame
results_df = pd.DataFrame(results)
print(results_df.head())

# Save results to Excel file (xlsx)
results_df.to_excel(output_filename, index=False)

# Open the Excel file to highlight the largest value in each row
wb = openpyxl.load_workbook(output_filename)
ws = wb.active

# Loop through rows and highlight the largest value in each row
for row in range(2, len(results_df) + 2):  # Start from row 2 because row 1 contains headers
    values = [ws.cell(row=row, column=col).value for col in range(3, len(results_df.columns) + 3)]  # Values from column 3 onwards
    
    # Filter out None values before applying max
    values = [value for value in values if value is not None]
    
    if values:  # Ensure there are values to compare
        max_value = min(values)
        max_index = values.index(max_value) + 3  # Adjust index for correct column in Excel
        
        # Bold the maximum value in the row
        ws.cell(row=row, column=max_index).font = Font(bold=True)

# Save the workbook with highlighted cells
wb.save(output_filename)
